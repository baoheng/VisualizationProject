import json

from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
from datetime import datetime


app = Flask(__name__)


@app.route('/')
@app.route('/index.html')
def index():
    print(getChart1Info())
    return render_template(
        'index.html',
        title='Home Page',
        chart1Info=getChart1Info(),
        chart2Info=getChart2Info(),
        year=datetime.now().year,
    )


@app.route('/chart1.html')
def chart1():
    return render_template(
        'chart1.html',
        title='Chart 1',
        chart1Info=getChart1Info()
    )


@app.route('/chart2.html')
def chart2():
    return render_template(
        'chart2.html',
        title='Chart 2',
        chart2Info=getChart2Info()
    )


@app.route('/createPieChart', methods=["POST"])
def createPieChart():
    sentInData = request.get_json(force=True)
    chart = None
    for key in sentInData.keys():
        chart = key

    if chart == "chart1":
        return jsonify(getPieChartInfo(sentInData[chart]["label"], getChart1Info()))
    elif chart == "chart2":
        return jsonify(getPieChartInfo(sentInData[chart]["label"], getChart2Info()))


def getPieChartInfo(year, chartInfo):
    oldData = chartInfo["data"]
    newData = []
    dic = {
        "type": "pie",
        "showInLegend": True,
        "toolTipContent": "{y} - #percent %",
        "yValueFormatString": "#0.#,,. Million",
        "legendText": "{indexLabel}",
    }
    dataPoints = {}
    for item in oldData:
        for dataItem in item["dataPoints"]:
            if dataItem["label"] == year:
                if item["name"] not in dataPoints.keys():
                    dataPoints[item["name"]] = {"y": dataItem["y"], "indexLabel": item["name"]}
                else:
                    dataPoints[item["name"]]["y"] += dataItem["y"]

    for item in dataPoints:
        newData.append(dataPoints[item])

    dic["dataPoints"] = newData
    pieChartInfo = {}
    pieChartInfo["title"] = "Outstanding Bonds and COPs FY in " + str(year)
    pieChartInfo["data"] = dic
    pieChartInfo["json"] = json.dumps(dic)
    return pieChartInfo


def getChart1Info():
    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb.get_sheet_by_name('Outstanding (Fig 2)')

    items = ["VP GO", "MVFT GO", "Triple Pledge", "GARVEEs", "TIFIA", "State COPs"]
    data = []

    for item in items:
        dict = {
            "type": "stackedColumn",
            "name": item,
            "cursor": "pointer",
            "showInLegend": True
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=3):
            if item == "VP GO":
                d = {
                    "label": row[0].value,
                    "y": row[1].value
                }
            elif item == "MVFT GO":
                d = {
                    "label": row[0].value,
                    "y": row[2].value
                }
            elif item == "Triple Pledge":
                d = {
                    "label": row[0].value,
                    "y": row[3].value
                }
            elif item == "GARVEEs":
                d = {
                    "label": row[0].value,
                    "y": row[4].value
                }
            elif item == "TIFIA":
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            else:
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            if d["label"] != None:
                dataPoints.append(d)

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = "Outstanding Bonds and COPs FY 2000-2016 ($ Billions)"
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)

    return chartInfo


def getChart2Info():
    wb = load_workbook(filename="static/Debt Affordability Study Data.xlsx", data_only=True)
    sheet = wb.get_sheet_by_name('New Money Issuance (Fig 3)')

    items = ["VP GO", "MVFT GO", "Triple Pledge", "GARVEEs", "TIFIA", "State COPs"]
    data = []

    for item in items:
        dict = {
            "type": "stackedColumn",
            "name": item,
            "cursor": "pointer",
            "showInLegend": True
        }

        dataPoints = []
        for row in sheet.iter_rows(row_offset=2):
            if item == "VP GO":
                d = {
                    "label": row[0].value,
                    "y": row[1].value
                }
            elif item == "MVFT GO":
                d = {
                    "label": row[0].value,
                    "y": row[2].value
                }
            elif item == "Triple Pledge":
                d = {
                    "label": row[0].value,
                    "y": row[3].value
                }
            elif item == "GARVEEs":
                d = {
                    "label": row[0].value,
                    "y": row[4].value
                }
            elif item == "TIFIA":
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            else:
                d = {
                    "label": row[0].value,
                    "y": row[5].value
                }
            if d["label"] != None:
                dataPoints.append(d)

        dict["dataPoints"] = dataPoints
        data.append(dict)

        # Create Chart Information
        chartInfo = {}
        chartInfo["chartTitle"] = "Bond and COP Issuance FY 2000-2016 ($ Millions)"
        chartInfo["data"] = data
        chartInfo["json"] = json.dumps(data)

    return chartInfo

if __name__ == '__main__':
    app.debug = True
    app.run()